from openpyxl import Workbook

wb = Workbook()

workSheet = wb.active

workSheet = wb.create_sheet('sheet1')

workSheet = wb['sheet1']

print(wb.active)
print(wb.sheetnames)